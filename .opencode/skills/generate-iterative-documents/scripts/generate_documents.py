#!/usr/bin/env python3
"""Render SRS/SD DOCX and populate STC XLSX from confirmed JSON input."""
import argparse
import json
import re
from collections import Counter
from pathlib import Path

from docxtpl import DocxTemplate
from openpyxl import load_workbook

CASE_TYPES = {
    "功能用例", "异常用例", "性能用例", "安全用例", "可靠性用例", "兼容性用例",
    "满规格", "时序", "资源回收", "并发", "边界校验",
}
PRIORITIES = {"H", "M", "L"}


def clean(value):
    return "" if value is None else str(value).strip()


def validate_data(data):
    """Reject structurally ambiguous input before rendering any document."""
    errors = []
    if not isinstance(data, dict):
        raise ValueError("input root must be a JSON object")

    for key in ("iteration_name", "iteration_code"):
        if not clean(data.get(key)):
            errors.append(f"{key}: required non-empty string")

    for section_name in ("srs", "sd"):
        section = data.get(section_name)
        if not isinstance(section, dict):
            errors.append(f"{section_name}: required object")

    requirements = data.get("requirements")
    if not isinstance(requirements, list) or not requirements:
        errors.append("requirements: required non-empty array")
        requirements = []
    for index, requirement in enumerate(requirements):
        path = f"requirements[{index}]"
        if not isinstance(requirement, dict):
            errors.append(f"{path}: must be an object")
            continue
        for key in ("title", "ar"):
            if not clean(requirement.get(key)):
                errors.append(f"{path}.{key}: required non-empty string")
        cases = requirement.get("test_cases", [])
        if not isinstance(cases, list):
            errors.append(f"{path}.test_cases: must be an array")
            continue
        for case_index, case in enumerate(cases):
            case_path = f"{path}.test_cases[{case_index}]"
            if not isinstance(case, dict):
                errors.append(f"{case_path}: must be an object")
                continue
            for key in ("id", "type", "item", "title", "priority", "precondition",
                        "input", "steps", "expected"):
                if not clean(case.get(key)):
                    errors.append(f"{case_path}.{key}: required non-empty string")
            case_type = clean(case.get("type"))
            if case_type and case_type not in CASE_TYPES:
                errors.append(f"{case_path}.type: unsupported value {case_type!r}")
            priority = clean(case.get("priority"))
            if priority and priority not in PRIORITIES:
                errors.append(f"{case_path}.priority: expected H, M, or L")

    revisions = data.get("revisions")
    if not isinstance(revisions, list):
        errors.append("revisions: required array")
    else:
        for index, revision in enumerate(revisions):
            path = f"revisions[{index}]"
            if not isinstance(revision, dict):
                errors.append(f"{path}: must be an object")
                continue
            for key in ("date", "version", "description", "author"):
                if not clean(revision.get(key)):
                    errors.append(f"{path}.{key}: required non-empty string")

    if errors:
        raise ValueError("invalid input:\n- " + "\n- ".join(errors))


def context(data, kind):
    iteration = clean(data.get("iteration_name"))
    code = clean(data.get("iteration_code")) or iteration.replace("迭代", "")
    requirements = data.get("requirements", [])
    revisions = data.get("revisions", [])
    if kind == "srs":
        section = data.get("srs", {})
        return {
            "HEADER": f"UnionSV110 {iteration}软件需求规格说明书",
            "TITLE_CN": f"UnionSV110 {iteration}需求软件规格说明书",
            "TITLE_EN": f"UnionSV110 {code} Iteration Software Requirement Specification",
            "PURPOSE": clean(section.get("purpose")),
            "SCOPE": clean(section.get("scope")),
            "OVERVIEW": clean(section.get("overview")),
            "SOFTWARE_FUNCTIONS": clean(section.get("software_functions")),
            "DESIGN_CONSTRAINTS": clean(section.get("design_constraints")),
            "PERFORMANCE_SUMMARY": clean(section.get("performance_summary")),
            "QUALITY_SUMMARY": clean(section.get("quality_summary")),
            "TESTABILITY_SUMMARY": clean(section.get("testability_summary")),
            "requirements": requirements,
            "revisions": revisions,
        }
    section = data.get("sd", {})
    return {
        "HEADER": f"UnionSV110 {iteration}需求软件设计说明书",
        "TITLE_CN": f"UnionSV110 {iteration}需求软件设计说明书",
        "TITLE_EN": f"UnionSV110 {code} Iterative Requirements Development Software Design Specification",
        "ITERATION_REQUIREMENT": f"{code}需求",
        "PURPOSE": clean(section.get("purpose")),
        "SCOPE": clean(section.get("scope")),
        "SOFTWARE_FUNCTIONS": clean(section.get("software_functions")),
        "designs": requirements,
        "revisions": revisions,
    }


def render_docx(template, output, values):
    doc = DocxTemplate(template)
    doc.render(values, autoescape=True)
    doc.save(output)


def safe_filename(value):
    return re.sub(r'[\\/:*?"<>|]+', "_", value).strip() or "iteration"


def populate_stc(template, output, data):
    wb = load_workbook(template)
    required = {"统计", "缺陷记录&测试报告"}
    missing = required.difference(wb.sheetnames)
    if missing:
        raise ValueError(f"STC template missing sheets: {', '.join(sorted(missing))}")
    ws = wb["缺陷记录&测试报告"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    cases = []
    for requirement in data.get("requirements", []):
        for case in requirement.get("test_cases", []):
            row = [
                clean(case.get("id")), clean(case.get("type")), clean(case.get("item")),
                clean(case.get("title")), clean(case.get("priority")), clean(case.get("precondition")),
                clean(case.get("input")), clean(case.get("steps")), clean(case.get("expected")),
                clean(case.get("execution")), clean(case.get("remarks")) or clean(requirement.get("ar")),
            ]
            if any(row):
                cases.append(row)
                ws.append(row)
    stats = wb["统计"]
    if stats.max_row:
        stats.delete_rows(1, stats.max_row)
    stats.append(["统计项", "数量"])
    stats.append(["用例总数", len(cases)])
    priorities = Counter(row[4] for row in cases)
    for priority in ("H", "M", "L"):
        stats.append([priority, priorities[priority]])
    types = Counter(row[1] for row in cases)
    for case_type in sorted(types):
        stats.append([case_type, types[case_type]])
    wb.save(output)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--srs-template", required=True, type=Path)
    parser.add_argument("--sd-template", required=True, type=Path)
    parser.add_argument("--stc-template", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()
    data = json.loads(args.input.read_text(encoding="utf-8"))
    validate_data(data)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    stem = safe_filename(clean(data.get("iteration_code")) or clean(data.get("iteration_name")))
    outputs = {
        "srs": args.output_dir / f"{stem}-SRS.docx",
        "sd": args.output_dir / f"{stem}-SD.docx",
        "stc": args.output_dir / f"{stem}-STC.xlsx",
    }
    render_docx(args.srs_template, outputs["srs"], context(data, "srs"))
    render_docx(args.sd_template, outputs["sd"], context(data, "sd"))
    populate_stc(args.stc_template, outputs["stc"], data)
    for path in outputs.values():
        print(path)


if __name__ == "__main__":
    main()
