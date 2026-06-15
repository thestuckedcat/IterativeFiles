import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


SCRIPT = Path(__file__).parents[1] / "scripts" / "generate_documents.py"
SPEC = importlib.util.spec_from_file_location("generate_documents", SCRIPT)
generate_documents = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(generate_documents)


class PopulateStcTests(unittest.TestCase):
    def test_writes_cases_to_test_case_sheet(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = root / "template.xlsx"
            output = root / "output.xlsx"
            workbook = Workbook()
            workbook.active.title = "统计"
            case_sheet = workbook.create_sheet("测试用例")
            case_sheet.append([
                "测试用例编号", "用例类型", "测试项", "测试用例标题", "重要级别",
                "预置条件", "输入", "操作步骤", "预期结果", "用例执行情况", "备注",
            ])
            report_sheet = workbook.create_sheet("缺陷记录&测试报告")
            report_sheet.append(["保留内容"])
            workbook.save(template)

            data = {
                "requirements": [{
                    "ar": "AR123",
                    "test_cases": [{
                        "id": "CASE-001",
                        "type": "功能用例",
                        "item": "导出",
                        "title": "导出测试用例",
                        "priority": "H",
                        "precondition": "无条件",
                        "input": "有效数据",
                        "steps": "执行导出",
                        "expected": "导出成功",
                    }],
                }],
            }

            generate_documents.populate_stc(template, output, data)

            result = load_workbook(output)
            self.assertEqual(result["测试用例"]["A2"].value, "CASE-001")
            self.assertEqual(result["测试用例"]["K2"].value, "AR123")
            self.assertEqual(result["缺陷记录&测试报告"]["A1"].value, "保留内容")
            self.assertEqual(result["统计"]["B2"].value, 1)

    def test_requires_test_case_sheet(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = root / "template.xlsx"
            workbook = Workbook()
            workbook.active.title = "统计"
            workbook.create_sheet("缺陷记录&测试报告")
            workbook.save(template)

            with self.assertRaisesRegex(ValueError, "测试用例"):
                generate_documents.populate_stc(template, root / "output.xlsx", {
                    "requirements": [],
                })


if __name__ == "__main__":
    unittest.main()
